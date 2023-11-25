from django.utils.html import format_html  # Import để định dạng HTML trong Django

import pandas as pd  # Import thư viện Pandas để làm việc với dữ liệu dạng bảng

from django.http import HttpResponse  # Import để tạo và trả về phản hồi HTTP trong Django

from openpyxl import Workbook  # Import để tạo và làm việc với tệp Excel trong Python sử dụng openpyxl

from openpyxl.styles import NamedStyle  # Import để tạo và sử dụng các kiểu định dạng trong openpyxl

from openpyxl.utils.dataframe import dataframe_to_rows  # Import để chuyển đổi dữ liệu từ DataFrame của Pandas thành hàng và cột trong openpyxl

import openpyxl  # Import thư viện openpyxl để làm việc với tệp Excel trong Python


# Tạo các mixin

class Mixins:
    def is_used_display(self, obj):
        return 'Use' if obj.is_used == 1 else 'Not'

    is_used_display.short_description = 'Is Used'
    
    def display_image(self, obj):
        if obj.image:
            image_url = obj.image.url
            return format_html('<img src="{}" style="max-width: 100px; max-height: 100px;" />', image_url)
        else:
            return "No image available"

    display_image.short_description = 'Image'

    def formatted_price(self, obj):
        price = obj.price
        formatted_price = f'{price}$'
        return formatted_price
    formatted_price.short_description = 'Price'

    def formatted_discount(self, obj):
        discount = obj.discount
        formatted_discount = f'{discount}%'
        return formatted_discount
    formatted_discount.short_description = 'Discount'

    def make_active(self, request, queryset):
        queryset.update(status='Active')
        for obj in queryset:
            obj.save()

    make_active.short_description = "Set selected brands as Active"

    def make_block(self, request, queryset):
        queryset.update(status='Block')
        for obj in queryset:
            obj.save()

    make_block.short_description = "Set selected brands as Block"

    def export_to_excel(modeladmin, request, queryset):
        # Lấy tên model để sử dụng làm tên file Excel
        model_name = modeladmin.model._meta.model_name

        # Chuyển đổi queryset thành DataFrame
        df = pd.DataFrame(list(queryset.values()))

        # Kiểm tra kiểu dữ liệu của cột 'created_at' và 'updated_at'
        print(df.dtypes)

        # Chuyển đổi cột 'created_at' và 'updated_at' sang dạng không có thông tin về múi giờ
        df['created_at'] = pd.to_datetime(df['created_at']).dt.tz_localize(None)
        df['updated_at'] = pd.to_datetime(df['updated_at']).dt.tz_localize(None)

        # Tạo response và thiết lập các thông số của file Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={model_name}_list.xlsx'

        # Tạo một workbook cho DataFrame
        workbook = Workbook()

        # Thêm một sheet vào workbook
        sheet = workbook.active
        sheet.title = "Sheet1"

        # Ghi DataFrame vào sheet
        for row_data in dataframe_to_rows(df, index=False, header=True):
            sheet.append(row_data)

        # Tạo một style cho cột 'created_at' và 'updated_at'
        date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD HH:MM:SS')
        

        # Tìm cột 'created_at' và 'updated_at' dựa trên tiêu đề
        created_at_col = None
        updated_at_col = None

        for col_num in range(1, sheet.max_column + 1):
            header_value = sheet.cell(row=1, column=col_num).value
            if header_value == 'created_at':
                created_at_col = col_num
            elif header_value == 'updated_at':
                updated_at_col = col_num

        # Áp dụng style chỉ nếu cả hai cột đều tồn tại
        if created_at_col is not None and updated_at_col is not None:
            for col_num in (created_at_col, updated_at_col):
                for row_num in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    cell.style = date_style
        else:
            print("Column 'created_at' or 'updated_at' was not found. Cannot apply style.")


        # Ghi workbook vào response
        workbook.save(response)

        return response

    export_to_excel.short_description = "Export selected Rows to Excel"


    